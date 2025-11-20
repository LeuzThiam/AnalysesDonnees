from __future__ import annotations

import os
import re
import json
import logging
import io
import base64
from datetime import datetime
import pandas as pd
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from rest_framework.decorators import api_view, parser_classes, permission_classes
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import AllowAny

# ✅ Imports internes
from .duck import (
    load_to_duckdb,
    list_tables,
    profile_table,
    run_sql,
    auto_analyze,
    query,
)
from .services.guards import is_safe
from .services.runners import run_sql_safe
from .services.planner import build_sql_from_plan
from integrations.n8n import nl_to_sql as n8n_nl_to_sql, is_configured as n8n_is_configured
from integrations.n8n_analysis import analyze_result, is_configured as analysis_is_configured

from .services.pandas_runner import run_pandas_analysis

# ============================================================
# 🔧 Détection si une question mérite un graphique
# ============================================================

def should_show_chart(question: str, chart_spec: dict, rows: list) -> bool:
    """
    Détermine si une question mérite un graphique ou juste une réponse textuelle.
    Retourne False pour les questions simples (liste, comptage, valeur unique).
    """
    q = question.lower().strip()
    
    # Si chart_spec existe et n'est pas "table", on affiche le graphique
    if chart_spec and chart_spec.get("type") and chart_spec.get("type") != "table":
        return True
    
    # Mots-clés indiquant qu'un graphique est approprié
    chart_keywords = [
        "évolution", "évolution", "courbe", "graphique", "visualiser", "afficher",
        "temps", "mois", "jour", "chronologique", "timeline", "par semaine", "par mois",
        "répartition", "proportion", "distribution", "part", "pourcentage",
        "classement", "top", "meilleur", "pire", "comparaison", "plus", "moins",
        "corrélation", "relation", "vs", "entre",
        "zone", "aire", "cumul", "progression",
        "tendance", "trend", "augmente", "diminue", "croissance", "baisse"
    ]
    
    # Mots-clés indiquant qu'un graphique n'est PAS nécessaire
    no_chart_keywords = [
        "liste", "lister", "afficher la liste", "donne la liste", "renvoie la liste",
        "combien", "nombre", "count", "total de", "nombre de",
        "quel est", "quelle est", "quelle valeur", "quelle est la valeur",
        "décris", "explique", "définis", "qu'est-ce que",
        "existe", "présent", "contient",
        "unique", "distinct", "différent"
    ]
    
    # Vérifier d'abord les mots-clés "pas de graphique"
    if any(keyword in q for keyword in no_chart_keywords):
        # Exception : si c'est "liste" mais avec "évolution" ou "comparaison", on garde le graphique
        if not any(chart_kw in q for chart_kw in ["évolution", "comparaison", "répartition", "classement"]):
            return False
    
    # Vérifier les mots-clés "graphique"
    if any(keyword in q for keyword in chart_keywords):
        return True
    
    # Si peu de lignes (≤ 5) et question simple, pas de graphique
    if rows and len(rows) <= 5:
        if any(simple in q for simple in ["liste", "combien", "quel est", "quelle est"]):
            return False
    
    # Par défaut, si on a des données avec plusieurs colonnes numériques, on peut faire un graphique
    if rows and len(rows) > 0:
        first_row = rows[0] if isinstance(rows[0], dict) else {}
        numeric_cols = [k for k, v in first_row.items() if isinstance(v, (int, float)) and v is not None]
        if len(numeric_cols) >= 1 and len(rows) > 1:
            return True
    
    # Par défaut, pas de graphique pour les questions simples
    return False


# ============================================================
# 🔧 Correction automatique du type de graphique (fallback)
# ============================================================

def auto_fix_chart_spec(question: str, chart_spec: dict, rows: list) -> dict | None:
    """Corrige ou déduit automatiquement le type de graphique selon la question.
    Retourne None si la question ne mérite pas de graphique."""
    
    # Vérifier d'abord si on doit afficher un graphique
    if not should_show_chart(question, chart_spec, rows):
        return None
    
    q = question.lower().strip()
    if not isinstance(chart_spec, dict):
        chart_spec = {}

    # 🧠 Détection du type si manquant ou 'table'
    if not chart_spec.get("type") or chart_spec.get("type") == "table":
        if any(k in q for k in ["répartition", "proportion", "distribution", "part", "pourcentage", "par nationalité", "par pays"]):
            chart_spec["type"] = "pie"
        elif any(k in q for k in ["évolution", "temps", "mois", "jour", "chronologique", "par semaine", "timeline"]):
            chart_spec["type"] = "line"
        elif any(k in q for k in ["classement", "top", "meilleur", "pire", "comparaison", "plus", "moins"]):
            chart_spec["type"] = "bar"
        elif any(k in q for k in ["corrélation", "relation", "vs", "entre"]):
            chart_spec["type"] = "scatter"
        elif any(k in q for k in ["zone", "aire", "cumul", "progression"]):
            chart_spec["type"] = "area"
        elif any(k in q for k in ["entonnoir", "funnel"]):
            chart_spec["type"] = "funnel"
        elif any(k in q for k in ["radial", "cercle concentrique", "progression circulaire"]):
            chart_spec["type"] = "radial_bar"
        elif any(k in q for k in ["treemap", "hiérarchie", "structure"]):
            chart_spec["type"] = "treemap"
        elif any(k in q for k in ["radar", "compétences", "profil", "polygone"]):
            chart_spec["type"] = "radar"
        elif any(k in q for k in ["empilé", "stacked"]):
            chart_spec["type"] = "stacked_bar"
        else:
            # Si aucun mot-clé détecté mais qu'on a décidé d'afficher un graphique, utiliser bar par défaut
            chart_spec["type"] = "bar"

    # 🔒 Sécurisation des clés x/y à partir des données réelles
    if "x" not in chart_spec and rows and isinstance(rows[0], dict):
        chart_spec["x"] = list(rows[0].keys())[0]
    if "y" not in chart_spec and rows and isinstance(rows[0], dict):
        chart_spec["y"] = list(rows[0].keys())[1] if len(rows[0]) > 1 else list(rows[0].keys())[0]

    return chart_spec


logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 🧩 Normalisation des noms
# ---------------------------------------------------------------------------

def _normalize_dataset_name(name: str) -> str:
    """Nettoie un nom de dataset pour qu’il soit compatible avec DuckDB."""
    return re.sub(r"[^A-Za-z0-9]+", "_", (name or "").strip()).strip("_").lower()


def _normalize_column_name(name: str) -> str:
    """Nettoie un nom de colonne pour l’usage SQL."""
    col = re.sub(r"\W+", "_", str(name).strip().lower())
    if re.match(r"^\d", col):
        col = f"col_{col}"
    return col


def _infer_left_var_from_read_csv(code: str) -> str | None:
    """Détecte la variable affectée au read_csv pour réinjection DuckDB."""
    if not code:
        return None
    m = re.search(r"^\s*([A-Za-z_]\w*)\s*=\s*pd\.read_csv\(", code, flags=re.MULTILINE)
    return m.group(1) if m else None


def _inject_duckdb_preamble(code: str, table_name: str, prefer_var: str | None = None) -> str:
    """Injecte le chargement DuckDB dans le code Python généré par le LLM."""
    left_var = _infer_left_var_from_read_csv(code) or prefer_var or "df"
    patched = re.sub(r"pd\.read_csv\([^)]*\)", left_var, code or "", flags=re.MULTILINE)

    preamble = f"""
import duckdb as _ddb
_con = _ddb.connect(r'{settings.DUCKDB_PATH}')
{left_var} = _con.execute('SELECT * FROM "{table_name}"').df()
"""
    return preamble.strip() + "\n\n" + patched.strip()


# ---------------------------------------------------------------------------
# 📊 Profilage automatique
# ---------------------------------------------------------------------------

_DATE_SYNONYMS = ("date", "jour", "day", "time", "timestamp", "datetime", "created", "due")
_NUM_SYNONYMS = ("cases", "cas", "total_cases", "value", "amount", "sum", "count", "nb", "y")
_ID_LIKE = {"id", "task id", "task_id"}

def _duck_profile(dataset: str) -> dict:
    return profile_table(dataset, limit=50) or {}


def _is_num_dtype(dt: str) -> bool:
    s = str(dt).lower()
    return any(k in s for k in ("int", "float", "double", "decimal", "numeric"))


def _is_dt_dtype(dt: str) -> bool:
    s = str(dt).lower()
    return any(k in s for k in ("date", "time", "timestamp", "datetime"))


def _infer_columns(dataset: str):
    """Retourne (date_col, num_col, cat_col) à partir du profil."""
    info = _duck_profile(dataset)
    cols = info.get("columns", [])

    # 🔹 Colonne de date
    date_col = next((c["name"] for c in cols if _is_dt_dtype(c["dtype"])), None)
    if not date_col:
        for p in _DATE_SYNONYMS:
            for c in cols:
                if p in str(c["name"]).lower().replace(" ", "_"):
                    date_col = c["name"]
                    break
            if date_col:
                break

    # 🔹 Colonne numérique
    num_col = next(
        (c["name"] for c in cols if _is_num_dtype(c["dtype"]) and str(c["name"]).lower() not in _ID_LIKE),
        None,
    )
    if not num_col:
        for p in _NUM_SYNONYMS:
            for c in cols:
                if p in str(c["name"]).lower().replace(" ", "_"):
                    num_col = c["name"]
                    break
            if num_col:
                break

    # 🔹 Colonne catégorielle
    cat_col = next(
        (c["name"] for c in cols if not _is_num_dtype(c["dtype"]) and not _is_dt_dtype(c["dtype"])),
        None,
    )

    return date_col, num_col, cat_col


# ---------------------------------------------------------------------------
# 🧠 Génération SQL automatique selon le type de graphique
# ---------------------------------------------------------------------------

def _synth_sql_from_spec(dataset: str, spec: dict) -> tuple[str | None, dict]:
    """Retourne (sql, spec_normalisée). Gère aussi les histogrammes automatiques."""
    if not isinstance(spec, dict):
        return None, spec or {}

    typ = (spec.get("type") or "").lower()
    spec_norm = dict(spec)

    date_col, num_col, cat_col = _infer_columns(dataset)

    def q(col: str) -> str:
        return f'"{col}"'

    # 🔹 Histogramme simple (ex : distribution des valeurs)
    if typ == "histogram":
        x = spec.get("x") or num_col
        if not x:
            return None, spec_norm
        spec_norm.setdefault("bins", 20)
        return f'SELECT {q(x)} FROM "{dataset}" WHERE {q(x)} IS NOT NULL', spec_norm

    # 🔹 Graphique en barres
    if typ in ("bar", "bar_horizontal"):
        x = spec.get("x") or cat_col
        y = spec.get("y") or num_col
        if not (x and y):
            return None, spec_norm
        sql = f'''
SELECT {q(x)} AS label, SUM(CAST({q(y)} AS DOUBLE)) AS value
FROM "{dataset}"
WHERE {q(x)} IS NOT NULL AND {q(y)} IS NOT NULL
GROUP BY 1
ORDER BY 2 DESC;
'''
        return sql.strip(), spec_norm

    # 🔹 Série temporelle
    if typ in ("line", "timeseries"):
        x = spec.get("x") or date_col
        y = spec.get("y") or num_col
        if not (x and y):
            return None, spec_norm
        sql = f'''
SELECT CAST({q(x)} AS TIMESTAMP) AS dt, SUM(CAST({q(y)} AS DOUBLE)) AS value
FROM "{dataset}"
WHERE {q(x)} IS NOT NULL AND {q(y)} IS NOT NULL
GROUP BY 1
ORDER BY 1;
'''
        return sql.strip(), spec_norm

    # 🔹 Graphique circulaire
    if typ == "pie":
        label = spec.get("label") or cat_col
        value = spec.get("value") or num_col
        if not (label and value):
            return None, spec_norm
        sql = f'''
SELECT {q(label)} AS name, SUM(CAST({q(value)} AS DOUBLE)) AS value
FROM "{dataset}"
WHERE {q(label)} IS NOT NULL AND {q(value)} IS NOT NULL
GROUP BY 1
ORDER BY 2 DESC;
'''
        return sql.strip(), spec_norm

    # 🔹 Table simple
    if typ == "table":
        return f'SELECT * FROM "{dataset}" LIMIT 1000;', spec_norm

    # 🧠 Nouvelle logique : détection automatique d’histogrammes
    # Si le LLM ne précise pas le type, on tente de deviner à partir du SQL
    if "GROUP BY" in spec.get("sql", "").upper() or "COUNT(" in spec.get("sql", "").upper():
        spec_norm.setdefault("type", "histogram")
        spec_norm.setdefault("x", cat_col or num_col or "x")
        spec_norm.setdefault("y", "count")

    return None, spec_norm


# ---------------------------------------------------------------------------
# 🌐 Endpoints API
# ---------------------------------------------------------------------------

@api_view(["POST"])
@parser_classes([MultiPartParser, FormParser])
@permission_classes([AllowAny])
def upload_dataset(request):
    """Upload et chargement dans DuckDB."""
    try:
        upfile = request.FILES.get("file")
        if not upfile:
            return JsonResponse({"detail": "Champ 'file' requis."}, status=400)

        dataset = _normalize_dataset_name(request.data.get("dataset") or os.path.splitext(upfile.name)[0])
        ext = (upfile.name or "").lower().rsplit(".", 1)[-1]

        if ext in ("csv", "xlsx", "xls", "json", "parquet"):
            info = load_to_duckdb(upfile, dataset, file_type=ext if ext != "xls" else "excel")
            return JsonResponse({"ok": True, "table": dataset, **info}, status=201)

        return JsonResponse({"detail": "Format non supporté (CSV, XLSX, JSON, Parquet)."}, status=400)

    except Exception as e:
        logger.exception("upload_dataset: erreur inattendue")
        return JsonResponse({"detail": f"Echec import: {e}"}, status=500)


@api_view(["GET"])
@permission_classes([AllowAny])
def datasets_list(request):
    try:
        return JsonResponse({"tables": list_tables()})
    except Exception as e:
        logger.exception("datasets_list: erreur inattendue")
        return JsonResponse({"detail": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([AllowAny])
def datasets_preview(request, table: str):
    """Preview limité d'un dataset (10-1000 lignes max)."""
    try:
        limit = max(1, min(int(request.GET.get("limit", 10)), 1000))
        info = profile_table(table, limit=limit)
        return JsonResponse({"table": table, **info})
    except Exception as e:
        logger.exception("datasets_preview: erreur inattendue")
        return JsonResponse({"detail": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([AllowAny])
def datasets_all(request, table: str):
    """
    Récupère TOUTES les données d'un dataset (sans limite).
    ⚠️ Attention : peut être très volumineux pour les gros datasets.
    Utilisez cette route uniquement si vous avez besoin de toutes les données.
    """
    try:
        dataset = _normalize_dataset_name(table)
        # Exécuter une requête SELECT * sans LIMIT
        sql = f'SELECT * FROM "{dataset}"'
        if not is_safe(sql):
            return JsonResponse({"detail": "Requête non autorisée."}, status=400)
        
        try:
            rows = run_sql_safe(sql, add_limit=None)  # Pas de limite
            return JsonResponse({
                "table": dataset,
                "rows": rows,
                "count": len(rows),
                "columns": list(rows[0].keys()) if rows else []
            })
        except Exception as e:
            logger.error(f"Erreur récupération données complètes ({dataset}): {e}")
            return JsonResponse({"detail": f"Erreur lors de la récupération des données: {e}"}, status=500)
    except Exception as e:
        logger.exception("datasets_all: erreur inattendue")
        return JsonResponse({"detail": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([AllowAny])
def query_sql(request):
    """Exécute une requête SQL brute (avec vérification de sécurité)."""
    try:
        sql = (request.data.get("sql") or "").strip()
        if not sql:
            return JsonResponse({"detail": "Champ 'sql' requis."}, status=400)

        if not is_safe(sql):
            return JsonResponse({"detail": "Requête SQL non autorisée."}, status=400)

        rows = run_sql_safe(sql)
        return JsonResponse({"rows": rows})
    except Exception as e:
        logger.exception("query_sql: erreur inattendue")
        return JsonResponse({"detail": str(e)}, status=500)


# ---------------------------------------------------------------------------
# 🧠 Requêtes NL → SQL via LLM
# ---------------------------------------------------------------------------

def get_schema(dataset: str) -> str:
    """Récupère les noms et types de colonnes DuckDB (pour contextualiser le LLM)."""
    if not dataset:
        return "Aucun dataset spécifié"

    try:
        df = query(f'DESCRIBE "{dataset}";')
        if df.empty:
            return "Aucune colonne détectée"
        cols = [f"{row['column_name']} ({row['column_type']})" for _, row in df.iterrows()]
        return ", ".join(cols)
    except Exception as e:
        logger.warning(f"Impossible de lire le schéma pour {dataset}: {e}")
        return "Schéma non disponible"
    

import numpy as np

def _format_sql_error(error_msg: str, sql: str = "") -> str:
    """
    Transforme une erreur SQL technique en message clair et compréhensible.
    """
    error_lower = error_msg.lower()
    
    # Erreur GROUP BY
    if "group by" in error_lower or "must appear in the group by" in error_lower:
        # Extraire le nom de la colonne - plusieurs patterns possibles
        col_name = None
        
        # Pattern 1: column "column_name" must appear
        col_match = re.search(r'column\s+"([^"]+)"\s+must appear', error_msg, re.IGNORECASE)
        if col_match:
            col_name = col_match.group(1)
        
        # Pattern 2: column column_name must appear (sans guillemets)
        if not col_name:
            col_match = re.search(r'column\s+([a-zA-Z_][a-zA-Z0-9_]*)\s+must appear', error_msg, re.IGNORECASE)
            if col_match:
                col_name = col_match.group(1)
        
        # Pattern 3: Binder Error avec nom de colonne
        if not col_name:
            col_match = re.search(r'Binder Error.*column\s+"([^"]+)"', error_msg, re.IGNORECASE)
            if col_match:
                col_name = col_match.group(1)
        
        # Pattern 4: Chercher dans le SQL fourni si disponible
        if not col_name and sql:
            # Chercher les colonnes dans le SELECT qui ne sont pas dans GROUP BY
            select_match = re.search(r'SELECT\s+(.+?)\s+FROM', sql, re.IGNORECASE | re.DOTALL)
            if select_match:
                select_cols = select_match.group(1)
                # Extraire les colonnes avec guillemets d'abord
                quoted_cols = re.findall(r'"([^"]+)"', select_cols)
                if quoted_cols:
                    # Vérifier si elles sont dans GROUP BY
                    group_by_match = re.search(r'GROUP\s+BY\s+(.+?)(?:\s+ORDER|\s+LIMIT|$)', sql, re.IGNORECASE | re.DOTALL)
                    if group_by_match:
                        group_by_cols = group_by_match.group(1)
                        # Trouver la colonne qui n'est pas dans GROUP BY
                        for col in quoted_cols:
                            if col.lower() not in group_by_cols.lower():
                                col_name = col
                                break
                    else:
                        # Pas de GROUP BY, prendre la première colonne non agrégée
                        col_name = quoted_cols[0] if quoted_cols else None
                
                # Si pas trouvé avec guillemets, chercher sans guillemets
                if not col_name:
                    # Extraire les noms de colonnes (sans fonctions d'agrégation)
                    simple_cols = re.findall(r'\b([a-zA-Z_][a-zA-Z0-9_]*)\b', select_cols)
                    # Filtrer les mots-clés SQL et fonctions d'agrégation
                    sql_keywords = {'select', 'from', 'where', 'group', 'by', 'order', 'limit', 'as', 'max', 'min', 'sum', 'avg', 'count', 'distinct', 'case', 'when', 'then', 'else', 'end'}
                    # Vérifier GROUP BY si présent
                    group_by_match = re.search(r'GROUP\s+BY\s+(.+?)(?:\s+ORDER|\s+LIMIT|$)', sql, re.IGNORECASE | re.DOTALL)
                    if group_by_match:
                        group_by_cols = group_by_match.group(1).lower()
                        # Trouver la colonne qui n'est pas dans GROUP BY
                        for col in simple_cols:
                            if col.lower() not in sql_keywords and col.lower() not in group_by_cols:
                                col_name = col
                                break
                    else:
                        # Pas de GROUP BY, prendre la première colonne non agrégée
                        col_name = next((col for col in simple_cols if col.lower() not in sql_keywords), None)
        
        if not col_name:
            col_name = "une colonne"
        
        return f"❌ Erreur dans la requête : La colonne '{col_name}' doit être incluse dans le GROUP BY ou utilisée dans une fonction d'agrégation (MAX, MIN, SUM, etc.).\n\n💡 Conseil : Reformulez votre question pour être plus précis, ou ajoutez '{col_name}' au GROUP BY."
    
    # Erreur de colonne inexistante
    if "does not exist" in error_lower or "column" in error_lower and "not found" in error_lower:
        col_match = re.search(r'column\s+"?([^"]+)"?\s+does not exist', error_msg, re.IGNORECASE)
        col_name = col_match.group(1) if col_match else "cette colonne"
        return f"❌ La colonne '{col_name}' n'existe pas dans ce dataset.\n\n💡 Conseil : Vérifiez le nom de la colonne ou consultez le schéma du dataset."
    
    # Erreur de table inexistante
    if "table" in error_lower and ("does not exist" in error_lower or "not found" in error_lower):
        table_match = re.search(r'table\s+"?([^"]+)"?\s+does not exist', error_msg, re.IGNORECASE)
        table_name = table_match.group(1) if table_match else "cette table"
        return f"❌ La table '{table_name}' n'existe pas.\n\n💡 Conseil : Vérifiez le nom du dataset ou importez-le d'abord."
    
    # Erreur de syntaxe
    if "syntax error" in error_lower or "invalid syntax" in error_lower:
        return f"❌ Erreur de syntaxe dans la requête SQL générée.\n\n💡 Conseil : Reformulez votre question de manière plus claire."
    
    # Erreur de type de données
    if "type mismatch" in error_lower or "cannot cast" in error_lower or "invalid type" in error_lower:
        return f"❌ Erreur de type de données : Les types de colonnes ne correspondent pas à l'opération demandée.\n\n💡 Conseil : Vérifiez que vous utilisez les bonnes colonnes pour votre analyse."
    
    # Erreur de fonction
    if "function" in error_lower and ("does not exist" in error_lower or "not found" in error_lower):
        func_match = re.search(r'function\s+"?([^"]+)"?\s+does not exist', error_msg, re.IGNORECASE)
        func_name = func_match.group(1) if func_match else "cette fonction"
        return f"❌ La fonction '{func_name}' n'est pas disponible ou n'existe pas.\n\n💡 Conseil : Utilisez une fonction SQL standard (SUM, COUNT, AVG, MAX, MIN, etc.)."
    
    # Erreur de division par zéro
    if "division by zero" in error_lower or "divide by zero" in error_lower:
        return f"❌ Division par zéro : Impossible de diviser par zéro.\n\n💡 Conseil : Vérifiez que les valeurs utilisées dans la division ne sont pas nulles."
    
    # Erreur de limite
    if "limit" in error_lower and ("invalid" in error_lower or "out of range" in error_lower):
        return f"❌ La limite spécifiée est invalide.\n\n💡 Conseil : Utilisez une valeur positive pour la limite."
    
    # Erreur de jointure
    if "join" in error_lower and ("ambiguous" in error_lower or "not found" in error_lower):
        return f"❌ Erreur dans la jointure de tables : Colonne ambiguë ou introuvable.\n\n💡 Conseil : Spécifiez explicitement les tables pour chaque colonne (ex: table.colonne)."
    
    # Erreur générique mais avec contexte
    if "binder error" in error_lower:
        # Essayer d'extraire des informations utiles
        if "group by" in error_lower:
            return f"❌ Erreur dans le GROUP BY : Toutes les colonnes non agrégées doivent être dans le GROUP BY.\n\n💡 Conseil : Reformulez votre question pour être plus précis sur les colonnes à regrouper."
        return f"❌ Erreur dans la requête SQL générée.\n\n💡 Conseil : Reformulez votre question de manière plus claire et précise."
    
    # Message générique avec extrait de l'erreur si elle est courte
    if len(error_msg) < 200:
        return f"❌ Erreur SQL : {error_msg}\n\n💡 Conseil : Reformulez votre question de manière plus claire."
    
    # Message générique pour les erreurs longues
    return f"❌ Erreur lors de l'exécution de la requête SQL.\n\n💡 Conseil : Reformulez votre question de manière plus claire et précise. Si le problème persiste, vérifiez que les noms de colonnes et de tables sont corrects."


def _format_text_response(question: str, rows: list[dict]) -> str:
    """
    Formate une réponse textuelle claire pour les questions qui ne méritent pas de graphique.
    """
    if not rows:
        return "Aucun résultat trouvé."
    
    q = question.lower().strip()
    num_rows = len(rows)
    first_row = rows[0] if rows else {}
    
    # Questions de comptage
    if any(kw in q for kw in ["combien", "nombre", "count", "total de", "nombre de", "combien de"]):
        if num_rows == 1 and isinstance(first_row, dict):
            # Si une seule ligne avec une valeur numérique
            numeric_values = {k: v for k, v in first_row.items() if isinstance(v, (int, float))}
            if numeric_values:
                key, value = next(iter(numeric_values.items()))
                return f"**{key}** : {value:,}" if isinstance(value, (int, float)) else f"**{key}** : {value}"
        return f"**Nombre de résultats** : {num_rows}"
    
    # Questions de liste simple
    if any(kw in q for kw in ["liste", "lister", "afficher la liste", "donne la liste", "renvoie la liste"]):
        if num_rows <= 10:
            # Afficher tous les éléments
            lines = []
            for i, row in enumerate(rows, 1):
                if isinstance(row, dict):
                    # Prendre les 2-3 premières colonnes significatives
                    cols = list(row.keys())[:3]
                    values = ", ".join(f"{k}: {v}" for k, v in row.items() if k in cols)
                    lines.append(f"{i}. {values}")
            return "\n".join(lines) if lines else "Aucun résultat."
        else:
            return f"**{num_rows} résultats trouvés.**\n\nAffichage des 10 premiers résultats dans le tableau ci-dessous."
    
    # Questions de valeur unique
    if any(kw in q for kw in ["quel est", "quelle est", "quelle valeur", "quelle est la valeur"]):
        if num_rows == 1 and isinstance(first_row, dict):
            # Afficher toutes les valeurs de la première ligne
            lines = []
            for k, v in first_row.items():
                if v is not None:
                    if isinstance(v, (int, float)):
                        lines.append(f"**{k}** : {v:,}")
                    else:
                        lines.append(f"**{k}** : {v}")
            return "\n".join(lines) if lines else "Aucune valeur trouvée."
        elif num_rows > 1:
            return f"**{num_rows} résultats trouvés.**\n\nVoir le tableau ci-dessous pour les détails."
    
    # Par défaut : résumé simple
    if num_rows == 1:
        if isinstance(first_row, dict):
            # Afficher les valeurs principales
            lines = []
            for k, v in list(first_row.items())[:5]:  # Limiter à 5 colonnes
                if v is not None:
                    if isinstance(v, (int, float)):
                        lines.append(f"**{k}** : {v:,}")
                    else:
                        lines.append(f"**{k}** : {v}")
            return "\n".join(lines) if lines else "Résultat trouvé."
    else:
        return f"**{num_rows} résultats trouvés.**\n\nVoir le tableau ci-dessous pour les détails."
    
    return f"{num_rows} résultat(s) trouvé(s)."


def auto_analyze_result(rows: list[dict], chart_spec: dict, question: str) -> str:
    """
    Génère une interprétation automatique simple selon les résultats.
    """
    if not rows or not isinstance(rows, list) or len(rows) == 0:
        return "Aucune donnée à analyser."

    # On essaie d’identifier les colonnes numériques et catégorielles
    first_row = rows[0]
    keys = list(first_row.keys())
    if len(keys) < 2:
        return "Résultat trop simple pour une analyse automatique."

    x_key = chart_spec.get("x", keys[0])
    y_key = chart_spec.get("y", keys[1])
    try:
        y_values = [float(r[y_key]) for r in rows if r.get(y_key) is not None]
    except Exception:
        return "Impossible d’interpréter les valeurs numériques."

    # Statistiques de base
    n = len(y_values)
    mean_val = np.mean(y_values)
    median_val = np.median(y_values)
    min_val = np.min(y_values)
    max_val = np.max(y_values)
    amplitude = max_val - min_val

    # Génération d’une analyse textuelle
    analysis = []

    # Type détecté
    typ = (chart_spec.get("type") or "").lower()

    if typ in ["line", "timeseries", "area"]:
        analysis.append(f"Les données présentent une évolution sur {n} points.")
        if y_values[-1] > y_values[0]:
            analysis.append("La tendance générale est à la hausse 📈.")
        elif y_values[-1] < y_values[0]:
            analysis.append("La tendance générale est à la baisse 📉.")
        else:
            analysis.append("La tendance reste stable sur la période.")
    elif typ in ["bar", "pie", "histogram", "stacked_bar"]:
        max_row = max(rows, key=lambda r: r[y_key])
        min_row = min(rows, key=lambda r: r[y_key])
        analysis.append(f"La catégorie '{max_row[x_key]}' a la valeur la plus élevée ({max_row[y_key]:.2f}).")
        analysis.append(f"La catégorie '{min_row[x_key]}' est la plus faible ({min_row[y_key]:.2f}).")
        if amplitude / mean_val > 0.5:
            analysis.append("La variation entre catégories est importante.")
        else:
            analysis.append("Les valeurs sont relativement homogènes entre catégories.")
    else:
        analysis.append("Les résultats sont affichés sous forme tabulaire. Consultez les valeurs clés ci-dessus.")

    analysis.append(f"Valeur moyenne : {mean_val:.2f}, médiane : {median_val:.2f}, min : {min_val:.2f}, max : {max_val:.2f}.")
    return " ".join(analysis)





@api_view(["POST"])
@permission_classes([AllowAny])
def query_nl(request):
    """
    NL → (n8n) → SQL/plan → exécution sécurisée → (optionnel) analyse experte n8n
    avec fallback local si n8n indisponible.
    """
    try:
        data = request.data or {}
        question = (data.get("question") or "").strip()
        dataset = _normalize_dataset_name(data.get("dataset"))

        if not question or not dataset:
            return JsonResponse({"detail": "Champs 'question' et 'dataset' requis."}, status=400)

        # 1) Schéma pour contextualiser
        schema = get_schema(dataset)
        extra = {k: v for k, v in data.items() if k not in {"question", "dataset"}}
        extra.update({"schema": schema})

        # 2) NL→SQL via n8n (si dispo)
        payload = {}
        if n8n_is_configured():
            try:
                payload = n8n_nl_to_sql(question, dataset, extra=extra)
            except Exception as e:
                logger.warning(f"Erreur n8n (NL→SQL): {e}")

        # 3) Cas code Python généré
        if payload.get("code_python"):
            code = _inject_duckdb_preamble(payload["code_python"], dataset, prefer_var=dataset)
            result = run_pandas_analysis(code)
            rows = result.get("rows", [])
            chart_spec = payload.get("chart_spec", {"type": "custom"})
            
            # Vérifier si on doit afficher un graphique
            chart_spec = auto_fix_chart_spec(question, chart_spec, rows)
            
            # Formater une réponse textuelle claire si pas de graphique
            text_response = None
            if not chart_spec and rows:
                text_response = _format_text_response(question, rows)
            
            analysis_text = ""
            analysis_summary = ""
            if analysis_is_configured():
                try:
                    # Envoyer toutes les données à n8n (rows contient déjà toutes les données)
                    logger.info(f"[query_nl] Envoi de {len(rows)} lignes à n8n pour analyse (dataset: {dataset}, code_python)")
                    n8n_out = analyze_result(question, rows, chart_spec)
                    raw_summary = n8n_out.get("summary") or n8n_out.get("text") or ""
                    
                    # Parser si c'est un JSON stringifié
                    try:
                        parsed = json.loads(raw_summary) if isinstance(raw_summary, str) else raw_summary
                        if isinstance(parsed, dict):
                            analysis_summary = parsed.get("summary", "")
                            analysis_text = parsed.get("text", "")
                        else:
                            analysis_text = raw_summary
                    except (json.JSONDecodeError, TypeError):
                        # Si ce n'est pas du JSON, utiliser tel quel
                        analysis_text = raw_summary
                except Exception as e:
                    logger.warning(f"Analyse n8n échouée: {e}")
            # Le summary contient seulement le summary de n8n NL→SQL (sans l'analyse experte)
            base_summary = payload.get("summary") or ""
            combined_summary = base_summary
            
            # Formater l'analyse experte pour l'affichage séparé
            formatted_analysis = ""
            if analysis_summary or analysis_text:
                if analysis_summary:
                    formatted_analysis = analysis_summary
                    if analysis_text and analysis_text != analysis_summary:
                        formatted_analysis += "\n\n" + analysis_text
                else:
                    formatted_analysis = analysis_text
            
            return JsonResponse({
                "rows": rows,
                "chart_spec": chart_spec,
                "summary": combined_summary,
                "analysis": formatted_analysis,
                "text_response": text_response,  # Réponse textuelle si pas de graphique
                "sql": payload.get("sql"),
                "schema": schema,
            })

        # 4) Cas SQL généré (ou synthèse depuis chart_spec / plan)
        chart_spec = payload.get("chart_spec", {}) or {}
        sql = (payload.get("sql") or "").strip()

        if not sql and chart_spec:
            sql, chart_spec = _synth_sql_from_spec(dataset, chart_spec)

        if not sql:
            date_col, val_col, cat_col = _infer_columns(dataset)
            plan = {
                "intent": data.get("intent", "timeseries_total"),
                "dataset": dataset,
                "date_col": date_col,
                "amount_col": val_col,
                "category_col": cat_col,
                "limit": int(data.get("limit", 1000)),
            }
            sql = build_sql_from_plan(plan)
            chart_spec = {"type": "table"}

        # 5) Sécurité puis exécution
        if not sql or not is_safe(sql):
            return JsonResponse({"detail": "SQL généré invalide ou non autorisé."}, status=400)

        try:
            # Exécuter sans limite pour avoir toutes les données pour n8n
            # On limite seulement pour l'affichage frontend si nécessaire
            rows = run_sql_safe(sql, add_limit=None)  # Pas de limite pour avoir toutes les données
        except Exception as e:
            logger.error(f"Erreur exécution SQL ({dataset}): {e}")
            # Formater l'erreur en message clair
            error_msg = str(e)
            # Extraire le message d'erreur original si c'est une QueryError ou RuntimeError
            if "Echec de l'exécution SQL:" in error_msg:
                # Extraire l'erreur originale après "Echec de l'exécution SQL:"
                original_error = error_msg.split("Echec de l'exécution SQL:", 1)[-1].strip()
                if "Erreur d'exécution SQL :" in original_error:
                    original_error = original_error.split("Erreur d'exécution SQL :", 1)[-1].strip()
                formatted_error = _format_sql_error(original_error, sql)
            else:
                formatted_error = _format_sql_error(error_msg, sql)
            return JsonResponse({"detail": formatted_error}, status=400)

        # 6) Fix chart + analyse locale / n8n
        chart_spec = auto_fix_chart_spec(question, chart_spec, rows)
        
        # Formater une réponse textuelle claire si pas de graphique
        text_response = None
        if not chart_spec and rows:
            text_response = _format_text_response(question, rows)

        analysis_text = ""
        analysis_summary = ""
        if analysis_is_configured():
            try:
                # Log pour vérifier le nombre de lignes envoyées
                logger.info(f"[query_nl] Envoi de {len(rows)} lignes à n8n pour analyse (dataset: {dataset})")
                n8n_out = analyze_result(question, rows, chart_spec)
                raw_summary = n8n_out.get("summary") or n8n_out.get("text") or ""
                
                # Parser si c'est un JSON stringifié
                try:
                    parsed = json.loads(raw_summary) if isinstance(raw_summary, str) else raw_summary
                    if isinstance(parsed, dict):
                        analysis_summary = parsed.get("summary", "")
                        analysis_text = parsed.get("text", "")
                    else:
                        analysis_text = raw_summary
                except (json.JSONDecodeError, TypeError):
                    # Si ce n'est pas du JSON, utiliser tel quel
                    analysis_text = raw_summary
            except Exception as e:
                logger.warning(f"Analyse n8n échouée: {e}")

        # Le summary contient seulement le summary de n8n NL→SQL (sans l'analyse experte)
        combined_summary = payload.get("summary") or ""

        # Formater l'analyse experte pour l'affichage séparé
        formatted_analysis = ""
        if analysis_summary or analysis_text:
            if analysis_summary:
                formatted_analysis = analysis_summary
                if analysis_text and analysis_text != analysis_summary:
                    formatted_analysis += "\n\n" + analysis_text
            else:
                formatted_analysis = analysis_text
        
        # Envoyer toutes les données (pas de limite)
        return JsonResponse({
            "rows": rows,  # Toutes les données, sans limite
            "chart_spec": chart_spec,
            "summary": combined_summary,
            "analysis": formatted_analysis,
            "text_response": text_response,  # Réponse textuelle si pas de graphique
            "sql": sql,
            "schema": schema,
        })

    except Exception as e:
        logger.exception("query_nl: erreur inattendue")
        return JsonResponse({"detail": f"Erreur interne: {e}"}, status=500)


# ============================================================
# 📥 EXPORT DES RÉSULTATS
# ============================================================

@api_view(["POST"])
@parser_classes([JSONParser])
@permission_classes([AllowAny])
def export_results(request):
    """
    Exporte les résultats d'une requête dans différents formats :
    - xlsx : Excel avec formatage
    - pdf : PDF avec graphiques et tableau
    - csv : CSV (déjà géré côté client, mais disponible ici aussi)
    """
    try:
        format_type = request.data.get("format", "xlsx").lower()
        rows = request.data.get("rows", [])
        question = request.data.get("question", "Résultats")
        dataset = request.data.get("dataset", "")
        chart_base64 = request.data.get("chart", None)  # Image base64 du graphique
        summary = request.data.get("summary", "")
        analysis = request.data.get("analysis", "")
        sql = request.data.get("sql", "")
        
        if not rows:
            return JsonResponse({"detail": "Aucune donnée à exporter."}, status=400)
        
        df = pd.DataFrame(rows)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"export_{dataset}_{timestamp}" if dataset else f"export_{timestamp}"
        
        if format_type == "xlsx":
            return _export_excel(df, question, dataset, summary, analysis, sql, chart_base64, filename_base)
        elif format_type == "pdf":
            return _export_pdf(df, question, dataset, summary, analysis, sql, chart_base64, filename_base)
        elif format_type == "csv":
            return _export_csv(df, filename_base)
        else:
            return JsonResponse({"detail": f"Format non supporté: {format_type}"}, status=400)
            
    except Exception as e:
        logger.exception("export_results: erreur inattendue")
        return JsonResponse({"detail": f"Erreur lors de l'export: {e}"}, status=500)


def _export_excel(df, question, dataset, summary, analysis, sql, chart_base64, filename_base):
    """Export Excel avec formatage et graphique."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.drawing.image import Image
        import tempfile
        
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        
        row_num = 1
        
        # Titre
        ws.merge_cells(f'A{row_num}:D{row_num}')
        ws[f'A{row_num}'] = f"Analyse : {question}"
        ws[f'A{row_num}'].font = title_font
        ws[f'A{row_num}'].alignment = Alignment(horizontal="center")
        row_num += 2
        
        # Informations
        if dataset:
            ws[f'A{row_num}'] = f"Dataset: {dataset}"
            row_num += 1
        if summary:
            ws[f'A{row_num}'] = "Résumé:"
            ws[f'A{row_num}'].font = Font(bold=True)
            row_num += 1
            ws[f'A{row_num}'] = summary
            ws.merge_cells(f'A{row_num}:D{row_num + summary.count(chr(10))}')
            row_num += summary.count(chr(10)) + 2
        if analysis:
            ws[f'A{row_num}'] = "Analyse experte:"
            ws[f'A{row_num}'].font = Font(bold=True)
            row_num += 1
            ws[f'A{row_num}'] = analysis
            ws.merge_cells(f'A{row_num}:D{row_num + analysis.count(chr(10))}')
            row_num += analysis.count(chr(10)) + 2
        
        # Graphique (si disponible)
        if chart_base64:
            try:
                img_data = base64.b64decode(chart_base64.split(',')[-1] if ',' in chart_base64 else chart_base64)
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                    tmp.write(img_data)
                    tmp_path = tmp.name
                
                img = Image(tmp_path)
                img.width = 600
                img.height = 400
                ws.add_image(img, f'A{row_num}')
                row_num += 25  # Espace pour l'image
                os.unlink(tmp_path)
            except Exception as e:
                logger.warning(f"Impossible d'ajouter le graphique à l'export Excel: {e}")
        
        row_num += 1
        
        # En-têtes
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row_num += 1
        
        # Données
        for _, row in df.iterrows():
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            row_num += 1
        
        # Ajuster la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return response
        
    except ImportError:
        logger.error("openpyxl n'est pas installé. Installation requise: pip install openpyxl")
        return JsonResponse({"detail": "Export Excel non disponible. Installez openpyxl: pip install openpyxl"}, status=500)
    except Exception as e:
        logger.exception("_export_excel: erreur")
        return JsonResponse({"detail": f"Erreur export Excel: {e}"}, status=500)


def _export_pdf(df, question, dataset, summary, analysis, sql, chart_base64, filename_base):
    """Export PDF avec graphique et tableau."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        import tempfile
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Style personnalisé pour le titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=1  # Centré
        )
        
        # Titre
        story.append(Paragraph(f"Analyse : {question}", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Informations
        if dataset:
            story.append(Paragraph(f"<b>Dataset:</b> {dataset}", styles['Normal']))
            story.append(Spacer(1, 0.1*inch))
        
        if summary:
            story.append(Paragraph("<b>Résumé:</b>", styles['Heading2']))
            story.append(Paragraph(summary.replace('\n', '<br/>'), styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        if analysis:
            story.append(Paragraph("<b>Analyse experte:</b>", styles['Heading2']))
            story.append(Paragraph(analysis.replace('\n', '<br/>'), styles['Normal']))
            story.append(Spacer(1, 0.2*inch))
        
        # Graphique (si disponible)
        if chart_base64:
            try:
                img_data = base64.b64decode(chart_base64.split(',')[-1] if ',' in chart_base64 else chart_base64)
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                    tmp.write(img_data)
                    tmp_path = tmp.name
                
                img = RLImage(tmp_path, width=6*inch, height=4*inch)
                story.append(img)
                story.append(Spacer(1, 0.2*inch))
                os.unlink(tmp_path)
            except Exception as e:
                logger.warning(f"Impossible d'ajouter le graphique à l'export PDF: {e}")
        
        # Tableau de données
        if len(df) > 0:
            story.append(Paragraph("<b>Données:</b>", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            
            # Préparer les données pour le tableau (limiter à 100 lignes pour le PDF)
            data = [df.columns.tolist()]  # En-têtes
            for _, row in df.head(100).iterrows():  # Limiter à 100 lignes pour le PDF
                data.append([str(val) for val in row.values])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            
            story.append(table)
            
            if len(df) > 100:
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph(f"<i>Note: Seules les 100 premières lignes sont affichées ({len(df)} lignes au total)</i>", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
        return response
        
    except ImportError:
        logger.error("reportlab n'est pas installé. Installation requise: pip install reportlab")
        return JsonResponse({"detail": "Export PDF non disponible. Installez reportlab: pip install reportlab"}, status=500)
    except Exception as e:
        logger.exception("_export_pdf: erreur")
        return JsonResponse({"detail": f"Erreur export PDF: {e}"}, status=500)


def _export_csv(df, filename_base):
    """Export CSV simple."""
    buffer = io.StringIO()
    df.to_csv(buffer, index=False, encoding='utf-8-sig')  # utf-8-sig pour Excel
    
    response = HttpResponse(buffer.getvalue(), content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    return response